"""Tests for Plan D batch 2 renderers — pdf-exec, xlsx-bukukerja, xlsx-generic."""
from datetime import date
from pathlib import Path

import pytest
from openpyxl import load_workbook

from pqcscan.core.types import Classification, Finding, Severity
from pqcscan.renderers.pdf_executive import render_pdf_executive
from pqcscan.renderers.xlsx_bukukerja import render_xlsx_bukukerja
from pqcscan.renderers.xlsx_generic import render_xlsx_generic
from pqcscan.store.repo import Repo


def _seed(repo: Repo) -> int:
    scan_id = repo.create_scan(mode="user", probe_versions={}, tool_versions={})
    fid_rsa = repo.record_finding(scan_id, Finding(
        probe_id="net.tls.https", algorithm="RSA-2048",
        classification=Classification.SANGAT_TINGGI, severity=Severity.CRIT,
        title="server cert uses RSA-2048",
        evidence={"endpoint": "127.0.0.1:443"},
    ))
    repo.record_finding(scan_id, Finding(
        probe_id="sbom.os.dpkg", algorithm="N/A",
        classification=Classification.INFO, severity=Severity.INFO,
        title="package: openssl 3.0.2-1",
        evidence={"name": "openssl", "version": "3.0.2-1",
                  "manager": "dpkg", "purl": "pkg:deb/openssl@3.0.2-1"},
    ))
    repo.record_framework_view(
        fid_rsa, framework="bukukerja",
        clause="BUKUKERJA:risk-register/sangat-tinggi",
        verdict="non-compliant", deadline=None,
    )
    repo.record_framework_view(
        fid_rsa, framework="cnsa2",
        clause="CNSA2:RSA-deprecated", verdict="non-compliant",
        deadline=date(2030, 12, 31),
    )
    repo.finish_scan(scan_id, status="done")
    return scan_id


def test_pdf_executive_writes_a_real_pdf(tmp_db_path, tmp_path: Path):
    pytest.importorskip("weasyprint")
    repo = Repo(tmp_db_path); repo.init_schema()
    scan_id = _seed(repo)
    out = tmp_path / "exec.pdf"
    render_pdf_executive(repo, scan_id, out)
    data = out.read_bytes()
    assert data.startswith(b"%PDF-")
    assert len(data) > 1000


def test_pdf_executive_raises_on_missing_scan(tmp_db_path, tmp_path: Path):
    pytest.importorskip("weasyprint")
    repo = Repo(tmp_db_path); repo.init_schema()
    with pytest.raises(ValueError):
        render_pdf_executive(repo, scan_id=999, output_path=tmp_path / "x.pdf")


def test_xlsx_generic_has_findings_sheet(tmp_db_path, tmp_path: Path):
    repo = Repo(tmp_db_path); repo.init_schema()
    scan_id = _seed(repo)
    out = tmp_path / "gen.xlsx"
    render_xlsx_generic(repo, scan_id, out)
    wb = load_workbook(out)
    assert wb.sheetnames == ["Findings"]
    ws = wb["Findings"]
    headers = [c.value for c in ws[1]]
    assert headers[:5] == ["id", "probe_id", "algorithm", "classification", "severity"]
    # Two findings + 1 header row.
    assert ws.max_row == 3


def test_xlsx_bukukerja_has_required_sheets(tmp_db_path, tmp_path: Path):
    repo = Repo(tmp_db_path); repo.init_schema()
    scan_id = _seed(repo)
    out = tmp_path / "bk.xlsx"
    render_xlsx_bukukerja(repo, scan_id, out)
    wb = load_workbook(out)
    expected = {"00_ReadMe", "0_Inventory", "1_SBOM", "2_CBOM", "3_RiskRegister"}
    assert expected.issubset(set(wb.sheetnames))


def test_xlsx_bukukerja_sbom_sheet_lists_packages(tmp_db_path, tmp_path: Path):
    repo = Repo(tmp_db_path); repo.init_schema()
    scan_id = _seed(repo)
    out = tmp_path / "bk.xlsx"
    render_xlsx_bukukerja(repo, scan_id, out)
    wb = load_workbook(out)
    sbom = wb["1_SBOM"]
    # Official Lampiran A template: title row 1, description row 2, blank row 3,
    # headers row 4, data rows 5+. Columns 1..17 follow the BUKUKERJA spec:
    #   1=# 2=System/Application 3=Purpose 4=URL 5=Services Mode 6=Target Customer
    #   7=Software Component ... 14=Vendor's Name ... 17=Link to CBOM
    rows = list(sbom.iter_rows(min_row=5, values_only=True))
    # Asset name col (col 2) is "openssl" (from evidence.name); component col
    # (col 7) is the combined "openssl 3.0.2-1" produced by the renderer.
    assert any(r[1] == "openssl" and r[6] == "openssl 3.0.2-1" for r in rows), (
        f"openssl row missing — got rows: {rows!r}"
    )


def test_xlsx_bukukerja_risk_register_carries_classification(tmp_db_path, tmp_path: Path):
    repo = Repo(tmp_db_path); repo.init_schema()
    scan_id = _seed(repo)
    out = tmp_path / "bk.xlsx"
    render_xlsx_bukukerja(repo, scan_id, out)
    wb = load_workbook(out)
    risk = wb["3_RiskRegister"]
    # Headers row 4, data rows 5+. Lampiran A Table 3 columns:
    #   1=# 2=Nama Sistem 3=Jenis Aset 4=Algoritma Kriptografi
    #   5=Kegunaan 6=Tahap Kritikal 7=Risiko 8=Pemilik Risiko
    rows = list(risk.iter_rows(min_row=5, values_only=True))
    assert any(
        r[3] == "RSA-2048" and r[5] == "sangat-tinggi"
        for r in rows
    ), f"sangat-tinggi RSA-2048 row missing — got rows: {rows!r}"
