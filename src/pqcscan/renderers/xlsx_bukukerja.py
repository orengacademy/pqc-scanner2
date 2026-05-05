"""renderers.xlsx_bukukerja — Malaysian BUKUKERJA workbook (NACSA Lampiran A).

Loads the bundled official template `templates/bukukerja-template.xlsx`
(8 sheets, including the static 5_RiskMatrix and 138-row 6_ProtocolCryptoMap
reference data) and populates the five dynamic sheets with rows derived
from the scan findings:

  0_Inventory      — Table 0: Initial Inventory for PQC Readiness (9 cols)
  1_SBOM           — Table 1: Software Bill of Materials (17 cols)
  2_CBOM           — Table 2: Cryptographic Bill of Materials (8 cols)
  3_RiskRegister   — Table 3: Risk Register (8 cols, BM headers)
  4_RiskAssessment — Table 4: Risk & Dependency Assessment (11 cols)

Static sheets `5_RiskMatrix`, `6_ProtocolCryptoMap`, `00_ReadMe` are left
intact so the workbook is a drop-in submission for NACSA Arahan KE No. 9
Borang Pelaksanaan Migrasi PQC. Headers in the template start at row 4;
data fills row 5+. We delete the example "Contoh:" rows shipped with the
template before populating.

Locale support: `locale="ms"` (default) emits Bahasa Malaysia per the
official template. `locale="en"` swaps every dynamic string + the BM
headers on sheets 3 / 4 to English equivalents so the same workbook is
usable by an English-speaking auditor without translating in-place.
"""
from __future__ import annotations

from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment

from pqcscan.store.repo import Repo
from pqcscan.store.schema import FindingRow

_TEMPLATE_PATH = Path(__file__).parent / "templates" / "bukukerja-template.xlsx"

# pqcscan classification → BUKUKERJA Risk Matrix Impact (1-5).
_IMPACT_BY_CLASSIFICATION: dict[str, int] = {
    "sangat-tinggi": 5,
    "tinggi": 4,
    "sederhana": 3,
    "rendah": 2,
    "info": 1,
    "error": 1,
    "pqc-ready": 1,
}

# Probe family → Likelihood (1-5).
_LIKELIHOOD_BY_FAMILY_PREFIX: dict[str, int] = {
    "net.": 5, "host.": 4, "vpn.": 4, "app.": 4,
    "storage.": 3, "fs.": 3, "container.": 3,
    "code.": 3, "sign.": 3, "dns_email.": 3,
    "sbom.": 2, "secrets.": 4, "aux.": 1, "pqc_meta.": 1,
}


def _likelihood_for(probe_id: str) -> int:
    pid = (probe_id or "").lower()
    for prefix, val in _LIKELIHOOD_BY_FAMILY_PREFIX.items():
        if pid.startswith(prefix):
            return val
    return 3


# ───── Bilingual label tables (ms / en) ─────
# Each tuple is (Bahasa Malaysia, English) — picked by `_pick(t, locale)`.

def _pick(t: tuple[str, str], locale: str) -> str:
    return t[1] if locale == "en" else t[0]


_JENIS_ASET: dict[str, tuple[str, str]] = {
    "net.":     ("Perkhidmatan Rangkaian",                 "Network Service"),
    "host.":    ("Sistem Pengendalian / Konfigurasi Hos",  "Operating System / Host Configuration"),
    "vpn.":     ("Perisian VPN",                           "VPN Software"),
    "storage.": ("Storan",                                 "Storage"),
    "fs.":      ("Sistem Fail / Sijil",                    "File System / Certificates"),
    "code.":    ("Kod Sumber",                             "Source Code"),
    "sbom.":    ("Pakej Perisian",                         "Software Package"),
    "container.": ("Kontena / Imej Kontena",               "Container / Container Image"),
    "app.":     ("Aplikasi",                               "Application"),
    "sign.":    ("Sijil / Tandatangan Digital",            "Certificate / Digital Signature"),
    "dns_email.": ("DNS / E-mel",                          "DNS / Email"),
    "secrets.": ("Rahsia",                                 "Secrets"),
    "aux.":     ("Tambahan",                               "Auxiliary"),
    "pqc_meta.": ("Metadata PQC",                          "PQC Metadata"),
    "trust.":   ("Stor Sijil Akar Sistem",                 "System Trust Root Store"),
}
_JENIS_ASET_DEFAULT: tuple[str, str] = ("Aset Lain", "Other Asset")


def _jenis_aset(probe_id: str, locale: str = "ms") -> str:
    pid = (probe_id or "").lower()
    for prefix, pair in _JENIS_ASET.items():
        if pid.startswith(prefix):
            return _pick(pair, locale)
    return _pick(_JENIS_ASET_DEFAULT, locale)


_KEGUNAAN: dict[str, tuple[str, str]] = {
    "tls":      ("Pertukaran Kunci & Pengesahan TLS",  "TLS Key Exchange & Authentication"),
    "ssh":      ("Pertukaran Kunci & Pengesahan SSH",  "SSH Key Exchange & Authentication"),
    "vpn":      ("Pertukaran Kunci VPN",               "VPN Key Exchange"),
    "cert":     ("Sijil & Tandatangan Digital",        "Certificates & Digital Signatures"),
    "sign":     ("Tandatangan Kod / Imej",             "Code / Image Signing"),
    "atrest":   ("Penyulitan At-Rest",                 "At-Rest Encryption"),
    "dnssec":   ("DNSSEC / DKIM",                      "DNSSEC / DKIM"),
    "lib":      ("Pelbagai (libraries)",               "Various (libraries)"),
    "kerberos": ("Pengesahan Kerberos",                "Kerberos Authentication"),
    "ldap":     ("Pengesahan LDAP",                    "LDAP Authentication"),
    "sig":      ("Tandatangan Digital",                "Digital Signature"),
    "kex":      ("Pertukaran Kunci",                   "Key Exchange"),
    "sym":      ("Penyulitan Simetri",                 "Symmetric Encryption"),
    "hash":     ("Hash / MAC",                         "Hash / MAC"),
    "any":      ("Pelbagai",                           "Various"),
}


def _kegunaan_kripto(probe_id: str, algorithm: str, locale: str = "ms") -> str:
    pid = (probe_id or "").lower()
    algo = (algorithm or "").upper()

    def t(key: str) -> str:
        return _pick(_KEGUNAAN[key], locale)

    if "tls" in pid or "starttls" in pid or "https" in pid:
        return t("tls")
    if pid.startswith("net.ssh") or ".ssh." in pid:
        return t("ssh")
    if pid.startswith("vpn.") or "wireguard" in pid or "openvpn" in pid:
        return t("vpn")
    if "cert" in pid or "x509" in pid or "trust" in pid:
        return t("cert")
    if pid.startswith("sign."):
        return t("sign")
    if pid.startswith(("storage.", "fs.fscrypt")):
        return t("atrest")
    if pid.startswith("dns_email."):
        return t("dnssec")
    if pid.startswith(("code.", "sbom.")):
        return t("lib")
    if "kerberos" in pid:
        return t("kerberos")
    if "ldap" in pid:
        return t("ldap")
    if any(x in algo for x in ("RSA", "DSA", "ECDSA", "ED25519", "ED448")):
        return t("sig")
    if any(x in algo for x in ("DH", "ECDH", "X25519", "X448", "ML-KEM", "KYBER")):
        return t("kex")
    if any(x in algo for x in ("AES", "CHACHA", "3DES", "DES", "RC4")):
        return t("sym")
    if any(x in algo for x in ("SHA", "MD5", "BLAKE", "SHA3")):
        return t("hash")
    return t("any")


# Mitigation / migration plan, indexed by algorithm fragment match.
_PELAN_MITIGASI_RULES: list[tuple[tuple[str, ...], tuple[str, str]]] = [
    (("ML-KEM", "MLKEM", "KYBER"),
     ("Sudah selari PQC (ML-KEM, FIPS 203). Pemantauan berterusan.",
      "Already PQC-aligned (ML-KEM, FIPS 203). Continuous monitoring.")),
    (("ML-DSA", "MLDSA", "DILITHIUM"),
     ("Sudah selari PQC (ML-DSA, FIPS 204). Pemantauan berterusan.",
      "Already PQC-aligned (ML-DSA, FIPS 204). Continuous monitoring.")),
    (("SLH-DSA", "SPHINCS"),
     ("Sudah selari PQC (SLH-DSA, FIPS 205). Pemantauan berterusan.",
      "Already PQC-aligned (SLH-DSA, FIPS 205). Continuous monitoring.")),
    (("RSA",),
     ("Migrasi kepada ML-KEM-768 (KEM, FIPS 203) atau ML-DSA-65 "
      "(signature, FIPS 204). Pertimbangkan hybrid X25519MLKEM768 "
      "untuk fasa peralihan.",
      "Migrate to ML-KEM-768 (KEM, FIPS 203) or ML-DSA-65 (signature, "
      "FIPS 204). Consider hybrid X25519MLKEM768 during transition.")),
    (("ECDSA", "ED25519", "ED448"),
     ("Migrasi kepada ML-DSA-65 (FIPS 204) atau SLH-DSA-128s "
      "(FIPS 205). Pertimbangkan hybrid signature untuk peralihan.",
      "Migrate to ML-DSA-65 (FIPS 204) or SLH-DSA-128s (FIPS 205). "
      "Consider hybrid signatures during transition.")),
    (("ECDH", "X25519", "X448"),
     ("Migrasi kepada ML-KEM-768 (FIPS 203) atau hybrid "
      "X25519MLKEM768 / P256MLKEM768.",
      "Migrate to ML-KEM-768 (FIPS 203) or hybrid X25519MLKEM768 / "
      "P256MLKEM768.")),
    (("DH-",),
     ("Migrasi kepada ML-KEM-768 (FIPS 203). DH klasik dimansuhkan.",
      "Migrate to ML-KEM-768 (FIPS 203). Classical DH is deprecated.")),
    (("AES-128", "AES128"),
     ("Naik taraf ke AES-256-GCM untuk hadapi serangan Grover "
      "(saiz kunci dua kali ganda).",
      "Upgrade to AES-256-GCM to withstand Grover (double the key size).")),
    (("3DES", "DES-", "RC4", "BLOWFISH"),
     ("Wajib gantikan dengan AES-256-GCM. Algoritma sudah dimansuhkan.",
      "Must replace with AES-256-GCM. Algorithm is deprecated.")),
    (("SHA-1", "SHA1"),
     ("Migrasi ke SHA-256 atau SHA-3 (FIPS 180-4 / FIPS 202). "
      "SHA-1 sudah dimansuhkan.",
      "Migrate to SHA-256 or SHA-3 (FIPS 180-4 / 202). SHA-1 is deprecated.")),
    (("MD5",),
     ("Migrasi ke SHA-256 atau SHA-3. MD5 dilarang sepenuhnya.",
      "Migrate to SHA-256 or SHA-3. MD5 is fully prohibited.")),
]
_PELAN_DEFAULT: tuple[str, str] = (
    "Rujuk Jadual 6 (Protocol Crypto Map) untuk algoritma pengganti PQC yang dicadangkan.",
    "Refer to Table 6 (Protocol Crypto Map) for the recommended PQC replacement algorithm.",
)
_PELAN_NO_ALGO: tuple[str, str] = (
    "Daftar aset dalam Borang Pelaksanaan Migrasi PQC (Lampiran A, Jadual 0).",
    "Register the asset in the PQC Migration Implementation Form (Appendix A, Table 0).",
)


def _pelan_mitigasi(algorithm: str, locale: str = "ms") -> str:
    algo = (algorithm or "").upper()
    if not algo or algo == "N/A":
        return _pick(_PELAN_NO_ALGO, locale)
    for fragments, plan in _PELAN_MITIGASI_RULES:
        if any(f in algo for f in fragments):
            return _pick(plan, locale)
    return _pick(_PELAN_DEFAULT, locale)


_PUNCA_RISIKO: dict[str, tuple[str, str]] = {
    "pqc":     ("Tiada — algoritma sudah selari PQC.",
                "None — algorithm is already PQC-aligned."),
    "shor":    ("Algoritma asimetri klasik — terdedah kepada serangan Shor pada CRQC.",
                "Classical asymmetric algorithm — vulnerable to Shor's attack on a CRQC."),
    "grover":  ("Algoritma simetri lemah — dimansuhkan atau dilemahkan oleh Grover.",
                "Weak symmetric algorithm — deprecated or weakened by Grover."),
    "hash":    ("Algoritma cincang dimansuhkan — risiko pelanggaran/pra-imej.",
                "Deprecated hash algorithm — collision/pre-image risk."),
    "sbom":    ("Algoritma dikunci dalam pakej perisian — keperluan kemaskini vendor.",
                "Algorithm is locked in a software package — vendor update required."),
    "code":    ("Algoritma terkunci dalam kod sumber — keperluan refactoring.",
                "Algorithm is locked in source code — refactoring required."),
    "cert":    ("Sijil X.509 menggunakan kunci klasik — keperluan rotasi sijil PQC.",
                "X.509 certificate uses classical keys — PQC certificate rotation required."),
    "default": ("Sistem/aplikasi sedia ada tidak menyokong algoritma PQC.",
                "Existing system/application does not support PQC algorithms."),
}


def _punca_risiko(probe_id: str, algorithm: str, locale: str = "ms") -> str:
    algo = (algorithm or "").upper()
    pid = (probe_id or "").lower()

    if any(x in algo for x in ("ML-KEM", "ML-DSA", "SLH-DSA", "MLKEM", "MLDSA")):
        return _pick(_PUNCA_RISIKO["pqc"], locale)
    if any(x in algo for x in ("RSA", "DSA", "ECDSA", "ECDH", "DH-", "X25519", "X448", "ED25519", "ED448")):
        return _pick(_PUNCA_RISIKO["shor"], locale)
    if any(x in algo for x in ("AES-128", "AES128", "3DES", "DES-", "RC4", "BLOWFISH")):
        return _pick(_PUNCA_RISIKO["grover"], locale)
    if any(x in algo for x in ("SHA-1", "SHA1", "MD5", "MD4")):
        return _pick(_PUNCA_RISIKO["hash"], locale)
    if pid.startswith("sbom."):
        return _pick(_PUNCA_RISIKO["sbom"], locale)
    if pid.startswith("code."):
        return _pick(_PUNCA_RISIKO["code"], locale)
    if "cert" in pid or "x509" in pid or "trust" in pid:
        return _pick(_PUNCA_RISIKO["cert"], locale)
    return _pick(_PUNCA_RISIKO["default"], locale)


_KAWALAN: dict[str, tuple[str, str]] = {
    "hybrid":   ("Hybrid PQC kex dikesan",      "Hybrid PQC kex detected"),
    "fips140":  ("Modul FIPS 140",              "FIPS 140 module"),
    "airgap":   ("Rangkaian terasing",          "Air-gapped network"),
}


def _kawalan_sedia_ada(finding: FindingRow, locale: str = "ms") -> str:
    ev = finding.evidence or {}
    controls: list[str] = []
    if any("MLKEM" in str(v).upper() or "KYBER" in str(v).upper() or "PQC" in str(v).upper()
           for v in ev.values()):
        controls.append(_pick(_KAWALAN["hybrid"], locale))
    if any("FIPS" in str(v).upper() for v in ev.values()):
        controls.append(_pick(_KAWALAN["fips140"], locale))
    if ev.get("network") == "isolated" or ev.get("airgap"):
        controls.append(_pick(_KAWALAN["airgap"], locale))
    return "; ".join(controls)


_RISK_LEVELS: list[tuple[int, tuple[str, str]]] = [
    (20, ("Risiko Sangat Tinggi", "Very High Risk")),
    (15, ("Risiko Tinggi",        "High Risk")),
    (10, ("Risiko Sederhana",     "Medium Risk")),
    (5,  ("Risiko Rendah",        "Low Risk")),
    (0,  ("Risiko Sangat Rendah", "Very Low Risk")),
]


def _risk_level_label(score: int, locale: str = "ms") -> str:
    for threshold, pair in _RISK_LEVELS:
        if score >= threshold:
            return _pick(pair, locale)
    return _pick(_RISK_LEVELS[-1][1], locale)


_MIGRATION_READINESS: dict[str, tuple[str, str]] = {
    "pqc-ready":     ("Tinggi",         "High"),
    "rendah":        ("Sederhana",      "Medium"),
    "sederhana":     ("Rendah",         "Low"),
    "tinggi":        ("Sangat Rendah",  "Very Low"),
    "sangat-tinggi": ("Sangat Rendah",  "Very Low"),
}
_READINESS_UNKNOWN: tuple[str, str] = ("Tidak Diketahui", "Unknown")


def _migration_readiness(classification: str, locale: str = "ms") -> str:
    pair = _MIGRATION_READINESS.get(classification)
    return _pick(pair if pair else _READINESS_UNKNOWN, locale)


def _asset_type_for(probe_id: str) -> str:
    return (probe_id or "").split(".", 1)[0].upper() or "UNKNOWN"


def _asset_name_for(finding: FindingRow) -> str:
    ev = finding.evidence or {}
    name = (
        ev.get("name") or ev.get("endpoint") or ev.get("host")
        or ev.get("path") or ev.get("dataset") or ev.get("device")
        or finding.probe_id
    )
    return str(name)


def _location_owner(finding: FindingRow) -> str:
    ev = finding.evidence or {}
    return (
        ev.get("location") or ev.get("owner") or ev.get("path")
        or ev.get("endpoint") or ""
    )


def _dedupe_risk(findings: list[FindingRow]) -> list[tuple[FindingRow, int]]:
    """Collapse near-identical risk-register rows."""
    groups: dict[tuple[str, str, str], list[FindingRow]] = defaultdict(list)
    for f in findings:
        key = (
            str(_asset_name_for(f)),
            _jenis_aset(f.probe_id),  # canonical BM key for grouping
            (f.algorithm or "").strip(),
        )
        groups[key].append(f)
    deduped: list[tuple[FindingRow, int]] = []
    for fs in groups.values():
        rep = max(fs, key=lambda f: _IMPACT_BY_CLASSIFICATION.get(str(f.classification), 0))
        deduped.append((rep, len(fs)))
    deduped.sort(key=lambda pair: (
        -_IMPACT_BY_CLASSIFICATION.get(str(pair[0].classification), 0),
        -pair[1], str(_asset_name_for(pair[0])),
    ))
    return deduped


def _delete_example_rows(ws, header_row: int = 4) -> None:
    while ws.max_row > header_row:
        ws.delete_rows(header_row + 1, ws.max_row - header_row)


def _wrap(cells) -> None:
    for c in cells:
        c.alignment = Alignment(wrap_text=True, vertical="top")


# Headers for sheets 3 and 4 — overridden when locale=en since the official
# template ships them in Bahasa Malaysia. Sheets 0/1/2 already have English
# headers in the template so they pass through unchanged.
_HEADERS_EN: dict[str, list[str]] = {
    "3_RiskRegister": [
        "#", "System / Hardware / Software Name",
        "Asset Type (Application/Hardware/Software)",
        "Cryptographic Algorithm", "Algorithm Usage",
        "Critical Level", "Risk", "Risk Owner",
    ],
    "4_RiskAssessment": [
        "#", "System / Hardware / Software Name",
        "Cryptographic Algorithm", "Risk", "Root Cause",
        "Impact", "Likelihood", "Risk Score", "Risk Level",
        "Existing Controls", "Mitigation Plan",
    ],
}


def _override_headers_en(ws, sheet_name: str) -> None:
    headers = _HEADERS_EN.get(sheet_name)
    if not headers:
        return
    for col_idx, label in enumerate(headers, start=1):
        if col_idx > ws.max_column:
            break
        ws.cell(row=4, column=col_idx).value = label


def render_xlsx_bukukerja(
    repo: Repo,
    scan_id: int,
    output_path: Path,
    *,
    locale: str = "ms",
) -> Path:
    """Render the BUKUKERJA workbook for `scan_id`.

    locale: "ms" (default, Bahasa Malaysia) or "en". When "en" is used,
    every dynamic string and the Bahasa headers on sheets 3/4 are
    swapped to English equivalents.
    """
    locale = "en" if locale == "en" else "ms"
    scan = repo.get_scan(scan_id)
    if scan is None:
        raise ValueError(f"scan {scan_id} not found")
    findings = repo.list_findings(scan_id)

    if not _TEMPLATE_PATH.is_file():
        raise FileNotFoundError(
            f"BUKUKERJA template missing at {_TEMPLATE_PATH}. "
            "Did the wheel build skip src/pqcscan/renderers/templates/?"
        )
    wb = load_workbook(_TEMPLATE_PATH)

    # ───────────  0_Inventory  ───────────
    inv = wb["0_Inventory"]
    _delete_example_rows(inv)
    grouped: dict[tuple[str, str], list[FindingRow]] = defaultdict(list)
    for f in findings:
        grouped[(_asset_type_for(f.probe_id), str(_asset_name_for(f)))].append(f)
    for idx, ((asset_type, asset_name), fs) in enumerate(sorted(grouped.items()), start=1):
        algos = sorted({f.algorithm for f in fs if f.algorithm and f.algorithm != "N/A"})
        sbom_present_pair = (("Ya", "Yes") if any(f.probe_id.startswith("sbom.") for f in fs)
                             else ("Tidak", "No"))
        worst = max(
            (str(f.classification) for f in fs),
            key=lambda c: _IMPACT_BY_CLASSIFICATION.get(c, 0),
            default="info",
        )
        notes = "; ".join(sorted({f.title for f in fs[:3]}))[:300]
        inv.append([
            idx,
            asset_type,
            asset_name[:120],
            _location_owner(fs[0])[:120],
            (("Ya", "Yes") if algos else ("Tidak", "No"))[1 if locale == "en" else 0],
            ", ".join(algos)[:200],
            sbom_present_pair[1 if locale == "en" else 0],
            _migration_readiness(worst, locale),
            notes,
        ])
        _wrap(inv[inv.max_row])

    # ───────────  1_SBOM  ───────────
    sbom_ws = wb["1_SBOM"]
    _delete_example_rows(sbom_ws)
    sbom_findings = [f for f in findings if f.probe_id.startswith("sbom.")]
    for idx, f in enumerate(sbom_findings, start=1):
        ev = f.evidence or {}
        component = ev.get("name") or ""
        if ev.get("version"):
            component = f"{component} {ev['version']}".strip()
        sbom_ws.append([
            idx, str(_asset_name_for(f))[:120],
            ev.get("manager", ""), ev.get("url", ""),
            "", "", component[:120], "", "", str(f.classification),
            "", "", "", ev.get("vendor", ""), "", "", "",
        ])
        _wrap(sbom_ws[sbom_ws.max_row])

    # ───────────  2_CBOM  ───────────
    cbom_ws = wb["2_CBOM"]
    _delete_example_rows(cbom_ws)
    cbom_findings = [
        f for f in findings
        if not f.probe_id.startswith("sbom.")
        and str(f.classification) != "info"
        and f.algorithm and f.algorithm != "N/A"
    ]
    for idx, f in enumerate(cbom_findings, start=1):
        ev = f.evidence or {}
        crypto_agility = ((("Ya", "Tidak Diketahui"), ("Yes", "Unknown"))[1 if locale == "en" else 0]
                          [0 if ev.get("crypto_agility") else 1])
        cbom_ws.append([
            f"CBOM #{idx}",
            str(_asset_name_for(f))[:120],
            f.probe_id.split(".", 2)[-1].replace("_", " ").title(),
            f.algorithm,
            ev.get("library", "") or ev.get("provider", ""),
            ev.get("key_size") or ev.get("key_length", ""),
            f.title[:200],
            crypto_agility,
        ])
        _wrap(cbom_ws[cbom_ws.max_row])

    # Risk-related sheets only emit findings classified high or above.
    risk_findings = [
        f for f in findings
        if str(f.classification) in ("sangat-tinggi", "tinggi", "sederhana")
    ]
    risk_grouped = _dedupe_risk(risk_findings)

    kejadian_word = "occurrences" if locale == "en" else "kejadian"

    def _suffix_count(text: str, count: int) -> str:
        if count <= 1:
            return text
        suffix = f"  (x{count} {kejadian_word})"
        return f"{text[:200 - len(suffix)]}{suffix}"

    # ───────────  3_RiskRegister  ───────────
    risk_ws = wb["3_RiskRegister"]
    _delete_example_rows(risk_ws)
    if locale == "en":
        _override_headers_en(risk_ws, "3_RiskRegister")
    for idx, (f, count) in enumerate(risk_grouped, start=1):
        risk_ws.append([
            idx,
            str(_asset_name_for(f))[:120],
            _jenis_aset(f.probe_id, locale),
            f.algorithm,
            _kegunaan_kripto(f.probe_id, f.algorithm, locale),
            str(f.classification),
            _suffix_count(f.title, count),
            "",
        ])
        _wrap(risk_ws[risk_ws.max_row])

    # ───────────  4_RiskAssessment  ───────────
    assess_ws = wb["4_RiskAssessment"]
    _delete_example_rows(assess_ws)
    if locale == "en":
        _override_headers_en(assess_ws, "4_RiskAssessment")
    for idx, (f, count) in enumerate(risk_grouped, start=1):
        impact = _IMPACT_BY_CLASSIFICATION.get(str(f.classification), 1)
        likelihood = _likelihood_for(f.probe_id)
        score = impact * likelihood
        assess_ws.append([
            idx,
            str(_asset_name_for(f))[:120],
            f.algorithm,
            _suffix_count(f.title, count),
            _punca_risiko(f.probe_id, f.algorithm, locale),
            impact, likelihood, score,
            _risk_level_label(score, locale),
            _kawalan_sedia_ada(f, locale),
            _pelan_mitigasi(f.algorithm, locale),
        ])
        _wrap(assess_ws[assess_ws.max_row])

    # 00_ReadMe, 5_RiskMatrix, 6_ProtocolCryptoMap left intact.
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    return output_path
