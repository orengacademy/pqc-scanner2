"""renderers.xlsx_generic — single-tab generic findings Excel sheet."""
from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from pqcscan.store.repo import Repo

_HEADER_FILL = PatternFill("solid", fgColor="0F172A")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADERS = (
    "id", "probe_id", "algorithm", "classification", "severity",
    "title", "evidence", "remediation", "created_at",
)


def render_xlsx_generic(repo: Repo, scan_id: int, output_path: Path) -> Path:
    scan = repo.get_scan(scan_id)
    if scan is None:
        raise ValueError(f"scan {scan_id} not found")
    findings = repo.list_findings(scan_id)

    wb = Workbook()
    ws = wb.active
    ws.title = "Findings"

    for col_idx, header in enumerate(_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")

    for row_idx, f in enumerate(findings, start=2):
        ws.cell(row=row_idx, column=1, value=f.id)
        ws.cell(row=row_idx, column=2, value=f.probe_id)
        ws.cell(row=row_idx, column=3, value=f.algorithm)
        ws.cell(row=row_idx, column=4, value=f.classification)
        ws.cell(row=row_idx, column=5, value=f.severity)
        ws.cell(row=row_idx, column=6, value=f.title)
        ws.cell(row=row_idx, column=7, value=json.dumps(f.evidence, default=str))
        ws.cell(row=row_idx, column=8, value=json.dumps(f.remediation, default=str))
        ws.cell(row=row_idx, column=9, value=f.created_at.isoformat())

    # Friendly column widths.
    widths = {1: 6, 2: 24, 3: 18, 4: 14, 5: 8, 6: 60, 7: 50, 8: 30, 9: 22}
    for col, w in widths.items():
        ws.column_dimensions[chr(64 + col)].width = w

    ws.freeze_panes = "A2"
    wb.save(str(output_path))
    return output_path
